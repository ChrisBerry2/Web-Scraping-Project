
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title


print(title.text)
##
##
##
##

table_rows = soup.findAll('tr')

wb = xl.Workbook()
write_sheet = wb.active

write_sheet.title = 'Box Office Movies'

write_sheet['A1'] = 'No.'
write_sheet['B1'] = 'Movie Title'
write_sheet['C1'] = 'Release Date'
write_sheet['D1'] = 'Number of Theaters'
write_sheet['E1'] = 'Total Gross'
write_sheet['F1'] = 'Average Gross by Theater'

for i in range(1,7):
    write_sheet.cell(1,1).font = Font(size = 16, bold = True)


write_sheet.column_dimensions['A'].width = 5
write_sheet.column_dimensions['B'].width = 30
write_sheet.column_dimensions['C'].width = 15
write_sheet.column_dimensions['D'].width = 20
write_sheet.column_dimensions['E'].width = 15
write_sheet.column_dimensions['F'].width = 25

headerfont = Font('Times New Roman', size = 12, italic=False, bold=True)

write_sheet['A1'] = 'No.'
write_sheet['B1'] = 'Movie Title'
write_sheet['C1'] = 'Release Date'
write_sheet['D1'] = 'Number of Theaters'
write_sheet['E1'] = 'Total Gross'
write_sheet['F1'] = 'Average gross by theater'

write_sheet['A1'].font = headerfont
write_sheet['B1'].font = headerfont
write_sheet['C1'].font = headerfont
write_sheet['D1'].font = headerfont
write_sheet['E1'].font = headerfont
write_sheet['F1'].font = headerfont

#find all table rows
table_rows = soup.find_all('tr')

#iterate over rows, extract data, and write to excel
for i in range(1, 6):
    data = table_rows[i].find_all('td')
    write_sheet.cell(row=i+1, column=1).value = i
    write_sheet.cell(row=i+1, column=2).value = data[0].text.strip()
    write_sheet.cell(row=i+1, column=3).value = data[1].text.strip()
    write_sheet.cell(row=i+1, column=4).value = data[2].text.strip()
    write_sheet.cell(row=i+1, column=5).value = data[3].text.strip()
    write_sheet.cell(row=i+1, column=6).value = data[4].text.strip()



wb.save('WebScrapedMovies.xlsx')